# apps/reportes/views.py
"""
Vistas del módulo de Reportes.

Reportes disponibles:
  - Estadísticas generales
  - Resultados por ficha
  - Resultados por trimestre
  - Resultados por competencia
   - Aprendices por estado (cumplen / pendientes / no cumplen)
  - Exportar a PDF (ReportLab fallback)
  - Exportar a Excel (openpyxl)
  - Historial de evaluaciones
"""

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.db.models import Count, Q, Avg, Sum, Case, When, IntegerField
from django.utils import timezone
from io import BytesIO

from apps.evaluacion.models import (
    Aprendiz, Evaluacion, EvaluacionItem,
    Resultado, Ficha, Fase, Competencia,
    Checklist, GAES, Trimestre,
)
from apps.usuarios.models import Usuario
import logging

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch
    from reportlab.lib import colors as rl_colors
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


# ─────────────────────────────────────────────────────────────────────────────
# Home / Estadísticas generales
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def home(request):
    if request.user.rol != 'administrador':
        return HttpResponseForbidden('Solo administradores pueden acceder a reportes')

    total_aprendices  = Aprendiz.objects.count()
    total_evaluaciones = Evaluacion.objects.count()
    total_resultados  = Resultado.objects.count()
    total_fichas      = Ficha.objects.count()
    total_competencias = Competencia.objects.count()
    total_checklists  = Checklist.objects.count()

    aprobados = Resultado.objects.filter(
        calificacion_final='Cumple'
    ).count()
    pendientes = Resultado.objects.filter(
        calificacion_final='No evaluado'
    ).count()
    no_cumplen  = Resultado.objects.exclude(
        calificacion_final__in=['Cumple', 'No evaluado']
    ).count()

    # Por ficha
    por_ficha = (
        Ficha.objects.annotate(
            cant_aprendices=Count('aprendices', distinct=True),
            cant_evaluaciones=Count('aprendices__evaluaciones', distinct=True),
            cant_aprobados=Count(
                'aprendices__resultados',
                filter=Q(aprendices__resultados__calificacion_final='Cumple'),
                distinct=True,
            ),
        )
        .order_by('numero')
    )

    # Por trimestre
    por_trimestre = (
        Trimestre.objects.annotate(
            cant_competencias=Count('competencias', distinct=True),
            cant_fichas=Count('fichas', distinct=True),
        )
        .order_by('-anio', 'numero')
    )

    # Por competencia
    por_competencia = (
        Competencia.objects.annotate(
            cant_eval=Count('items__evaluacionitem__evaluacion', distinct=True),
        )
        .order_by('codigo')
    )

    return render(request, 'reportes/home.html', {
        'total_aprendices': total_aprendices,
        'total_evaluaciones': total_evaluaciones,
        'total_resultados': total_resultados,
        'total_fichas': total_fichas,
        'total_competencias': total_competencias,
        'total_checklists': total_checklists,
        'aprobados': aprobados,
        'pendientes': pendientes,
        'no_cumplen': no_cumplen,
        'por_ficha': por_ficha,
        'por_trimestre': por_trimestre,
        'por_competencia': por_competencia,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Aprendices por estado
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def aprendices_estado(request, estado):
    if request.user.rol != 'administrador':
        return HttpResponseForbidden()

    if estado == 'cumplen':
        resultados = Resultado.objects.filter(
            calificacion_final='Cumple'
        ).select_related('aprendiz__ficha', 'aprendiz__gaes')
        titulo = 'Aprendices que Cumplen'
    elif estado == 'pendientes':
        resultados = Resultado.objects.filter(
            calificacion_final='No evaluado'
        ).select_related('aprendiz__ficha', 'aprendiz__gaes')
        titulo = 'Aprendices Pendientes'
    else:
        resultados = Resultado.objects.exclude(
            calificacion_final__in=['Cumple', 'No evaluado']
        ).select_related('aprendiz__ficha', 'aprendiz__gaes')
        titulo = 'Aprendices que No Cumplen'

    return render(request, 'reportes/aprendices_estado.html', {
        'titulo': titulo,
        'resultados': resultados,
        'estado': estado,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Resultados por ficha
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def resultados_por_ficha(request, ficha_id):
    if request.user.rol != 'administrador':
        return HttpResponseForbidden()

    ficha = get_object_or_404(Ficha, id=ficha_id)
    resultados = Resultado.objects.filter(
        aprendiz__ficha=ficha
    ).select_related('aprendiz').order_by('aprendiz__nombres')

    return render(request, 'reportes/resultados_ficha.html', {
        'ficha': ficha,
        'resultados': resultados,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Resultados por trimestre
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def resultados_por_trimestre(request, trimestre_id):
    if request.user.rol != 'administrador':
        return HttpResponseForbidden()

    trimestre = get_object_or_404(Trimestre, id=trimestre_id)
    resultados = Resultado.objects.filter(
        aprendiz__trimestre=str(trimestre.numero)
    ).select_related('aprendiz').order_by('aprendiz__nombres')

    return render(request, 'reportes/resultados_trimestre.html', {
        'trimestre': trimestre,
        'resultados': resultados,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Resultados por competencia
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def resultados_por_competencia(request, competencia_id):
    if request.user.rol != 'administrador':
        return HttpResponseForbidden()

    competencia = get_object_or_404(Competencia, id=competencia_id)
    items = ChecklistItem.objects.filter(
        competencia=competencia
    ).select_related('checklist').order_by('orden')

    return render(request, 'reportes/resultados_competencia.html', {
        'competencia': competencia,
        'items': items,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Historial de evaluaciones
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def historial_evaluaciones(request):
    if request.user.rol != 'administrador':
        return HttpResponseForbidden()

    evaluaciones = Evaluacion.objects.select_related(
        'aprendiz', 'juror', 'checklist'
    ).order_by('-fecha')

    search = request.GET.get('search', '')
    if search:
        evaluaciones = evaluaciones.filter(
            Q(aprendiz__nombres__icontains=search) |
            Q(aprendiz__apellidos__icontains=search) |
            Q(juror__username__icontains=search) |
            Q(checklist__titulo__icontains=search)
        )

    return render(request, 'reportes/historial_evaluaciones.html', {
        'evaluaciones': evaluaciones,
        'search': search,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Exportar PDF (ReportLab)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def exportar_pdf(request, tipo):
    """Genera un PDF con el resumen de resultados.

    tipo ∈ {aprendices, fichas, competencias, historial}
    """
    if request.user.rol != 'administrador':
        return HttpResponseForbidden()
    if not HAS_REPORTLAB:
        messages.error(request, 'ReportLab no esta instalado')
        return redirect('reportes_home')

    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch
    from reportlab.lib import colors as rl_colors

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    w, h = letter

    def encabezado(titulo):
        logo_path = settings.BASE_DIR / 'static' / 'images' / 'Logotipo_SENA.png'
        if logo_path.exists():
            c.drawImage(str(logo_path), 0.5 * inch, h - 0.95 * inch, width=0.6 * inch, height=0.6 * inch, preserveAspectRatio=True)
        c.setFont('Helvetica-Bold', 13)
        c.drawCentredString(w / 2, h - 0.6 * inch, titulo.upper())
        c.setFont('Helvetica', 8)
        c.drawCentredString(w / 2, h - 0.78 * inch,
                            'Sistema de Gestion de Sustentaciones SENA')
        c.setStrokeColorRGB(0.18, 0.25, 0.34)
        c.setLineWidth(1.5)
        c.line(0.75 * inch, h - 0.88 * inch, w - 0.75 * inch, h - 0.88 * inch)

    y = h - 1.1 * inch

    def nueva_pagina(titulo):
        nonlocal y
        c.showPage()
        encabezado(titulo)
        y = h - 1.1 * inch

    def ensure(n=1):
        nonlocal y
        if y < 0.85 * inch + n * 0.22 * inch:
            nueva_pagina(c._doc.title if hasattr(c, '_doc') else '')

    if tipo == 'aprendices':
        encabezado('Reporte de Aprendices')
        c.setFont('Helvetica-Bold', 11)
        c.drawString(0.9 * inch, y, 'Todos los aprendices registrados')
        y -= 0.3 * inch

        headers = ['Documento', 'Nombres', 'Apellidos', 'Programa', 'Ficha', 'GAES']
        col_widths = [1.1, 1.4, 1.4, 1.5, 0.9, 0.9]
        col_x = [0.9 * inch]
        for w_ in col_widths[:-1]:
            col_x.append(col_x[-1] + w_ * inch)

        c.setFillColorRGB(0.29, 0.44, 0.65)
        for i, h_text in enumerate(headers):
            c.rect(col_x[i], y - 2, col_widths[i] * inch, 0.22 * inch - 2, fill=1, stroke=0)
            c.setFillColorRGB(1, 1, 1)
            c.setFont('Helvetica-Bold', 7.5)
            c.drawString(col_x[i] + 3, y + 4, h_text)
        c.setFillColorRGB(0, 0, 0)
        y -= 0.22 * inch

        for i, ap in enumerate(Aprendiz.objects.select_related('ficha', 'gaes').order_by('nombres')):
            if y < 0.85 * inch:
                nueva_pagina('Reporte de Aprendices')
            bg = rl_colors.Color(0.97, 0.97, 0.97) if i % 2 == 0 else rl_colors.white
            vals = [
                ap.documento[:14],
                (ap.nombres or '')[:22],
                (ap.apellidos or '')[:22],
                (ap.programa or '')[:22],
                str(ap.ficha) if ap.ficha else '',
                str(ap.gaes) if ap.gaes else '',
            ]
            for j, v in enumerate(vals):
                c.setFillColorRGB(0.97, 0.97, 0.97) if i % 2 == 0 else c.setFillColorRGB(1, 1, 1)
                c.rect(col_x[j], y - 2, col_widths[j] * inch, 0.22 * inch - 2, fill=1, stroke=0)
                c.setFillColorRGB(0, 0, 0)
                c.setFont('Helvetica', 7.5)
                c.drawString(col_x[j] + 3, y + 3, v)
            y -= 0.22 * inch

    elif tipo == 'fichas':
        encabezado('Reporte de Fichas')
        c.setFont('Helvetica-Bold', 11)
        c.drawString(0.9 * inch, y, 'Resumen por ficha')
        y -= 0.3 * inch
        headers = ['Numero', 'Programa', 'GAES', 'Aprendices', 'Evaluaciones']
        col_widths = [1.0, 2.0, 1.2, 1.0, 1.0]
        col_x = [0.9 * inch]
        for w_ in col_widths[:-1]:
            col_x.append(col_x[-1] + w_ * inch)

        c.setFillColorRGB(0.29, 0.44, 0.65)
        for i, h_text in enumerate(headers):
            c.rect(col_x[i], y - 2, col_widths[i] * inch, 0.22 * inch - 2, fill=1, stroke=0)
            c.setFillColorRGB(1, 1, 1)
            c.setFont('Helvetica-Bold', 7.5)
            c.drawString(col_x[i] + 3, y + 4, h_text)
        c.setFillColorRGB(0, 0, 0)
        y -= 0.22 * inch

        for i, f in enumerate(Ficha.objects.select_related('gaes').prefetch_related('aprendices').order_by('numero')):
            if y < 0.85 * inch:
                nueva_pagina('Reporte de Fichas')
            cant_ap = f.aprendices.count()
            cant_ev = sum(
                a.evaluaciones.count() for a in f.aprendices.all()
            )
            vals = [f.numero[:16], (f.programa or '')[:30], str(f.gaes or ''),
                    str(cant_ap), str(cant_ev)]
            for j, v in enumerate(vals):
                c.setFillColorRGB(0.97, 0.97, 0.97) if i % 2 == 0 else c.setFillColorRGB(1, 1, 1)
                c.rect(col_x[j], y - 2, col_widths[j] * inch, 0.22 * inch - 2, fill=1, stroke=0)
                c.setFillColorRGB(0, 0, 0)
                c.setFont('Helvetica', 7.5)
                c.drawString(col_x[j] + 3, y + 3, v)
            y -= 0.22 * inch

    elif tipo == 'competencias':
        encabezado('Reporte de Competencias')
        c.setFont('Helvetica-Bold', 11)
        c.drawString(0.9 * inch, y, 'Resumen por competencia')
        y -= 0.3 * inch
        headers = ['Codigo', 'Nombre', 'Fase', 'Evaluaciones']
        col_widths = [1.2, 2.8, 0.8, 1.2]
        col_x = [0.9 * inch]
        for w_ in col_widths[:-1]:
            col_x.append(col_x[-1] + w_ * inch)

        c.setFillColorRGB(0.29, 0.44, 0.65)
        for i, h_text in enumerate(headers):
            c.rect(col_x[i], y - 2, col_widths[i] * inch, 0.22 * inch - 2, fill=1, stroke=0)
            c.setFillColorRGB(1, 1, 1)
            c.setFont('Helvetica-Bold', 7.5)
            c.drawString(col_x[i] + 3, y + 4, h_text)
        c.setFillColorRGB(0, 0, 0)
        y -= 0.22 * inch

        for i, comp in enumerate(Competencia.objects.select_related('fase').order_by('codigo')):
            if y < 0.85 * inch:
                nueva_pagina('Reporte de Competencias')
            cant_eval = EvaluacionItem.objects.filter(
                item__competencia=comp
            ).count()
            vals = [comp.codigo[:18], (comp.nombre or '')[:40],
                    str(comp.fase) if comp.fase else '', str(cant_eval)]
            for j, v in enumerate(vals):
                c.setFillColorRGB(0.97, 0.97, 0.97) if i % 2 == 0 else c.setFillColorRGB(1, 1, 1)
                c.rect(col_x[j], y - 2, col_widths[j] * inch, 0.22 * inch - 2, fill=1, stroke=0)
                c.setFillColorRGB(0, 0, 0)
                c.setFont('Helvetica', 7.5)
                c.drawString(col_x[j] + 3, y + 3, v)
            y -= 0.22 * inch

    elif tipo == 'historial':
        encabezado('Historial de Evaluaciones')
        c.setFont('Helvetica-Bold', 11)
        c.drawString(0.9 * inch, y, 'Todas las evaluaciones registradas')
        y -= 0.3 * inch
        headers = ['Fecha', 'Aprendiz', 'Jurado', 'Checklist', 'Estado']
        col_widths = [1.2, 1.6, 1.2, 1.8, 1.0]
        col_x = [0.9 * inch]
        for w_ in col_widths[:-1]:
            col_x.append(col_x[-1] + w_ * inch)

        c.setFillColorRGB(0.29, 0.44, 0.65)
        for i, h_text in enumerate(headers):
            c.rect(col_x[i], y - 2, col_widths[i] * inch, 0.22 * inch - 2, fill=1, stroke=0)
            c.setFillColorRGB(1, 1, 1)
            c.setFont('Helvetica-Bold', 7.5)
            c.drawString(col_x[i] + 3, y + 4, h_text)
        c.setFillColorRGB(0, 0, 0)
        y -= 0.22 * inch

        for i, ev in enumerate(Evaluacion.objects.select_related(
                'aprendiz', 'juror', 'checklist').order_by('-fecha')):
            if y < 0.85 * inch:
                nueva_pagina('Historial de Evaluaciones')
            vals = [
                ev.fecha.strftime('%d/%m/%Y') if ev.fecha else '',
                f'{ev.aprendiz.nombres} {ev.aprendiz.apellidos}'[:28],
                str(ev.juror.username or '—')[:20],
                (ev.checklist.titulo or '')[:30],
                ev.get_estado_display(),
            ]
            for j, v in enumerate(vals):
                c.setFillColorRGB(0.97, 0.97, 0.97) if i % 2 == 0 else c.setFillColorRGB(1, 1, 1)
                c.rect(col_x[j], y - 2, col_widths[j] * inch, 0.22 * inch - 2, fill=1, stroke=0)
                c.setFillColorRGB(0, 0, 0)
                c.setFont('Helvetica', 7.5)
                c.drawString(col_x[j] + 3, y + 3, v)
            y -= 0.22 * inch

    else:
        encabezado('Reporte')
        c.setFont('Helvetica', 10)
        c.drawString(0.9 * inch, y, f'Tipo de reporte no reconocido: {tipo}')

    # Footer
    if y > 0.55 * inch:
        y -= 0.15 * inch
    c.setStrokeColorRGB(0.7, 0.7, 0.7)
    c.setLineWidth(0.5)
    c.line(0.9 * inch, y + 0.1 * inch, w - 0.75 * inch, y + 0.1 * inch)
    c.setFont('Helvetica-Oblique', 8)
    c.setFillColorRGB(0.5, 0.5, 0.5)
    c.drawCentredString(w / 2, 0.55 * inch,
                        'Generado por Sistema de Gestion de Sustentaciones SENA')
    c.showPage()
    c.save()
    buf.seek(0)
    filename = f'reporte_{tipo}_{timezone.now().strftime("%Y%m%d")}.pdf'
    r = HttpResponse(buf, content_type='application/pdf')
    r['Content-Disposition'] = f'attachment; filename="{filename}"'
    return r


# ─────────────────────────────────────────────────────────────────────────────
# Exportar a Excel
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def exportar_excel(request, tipo):
    """Exporta un tipo de reporte a Excel (.xlsx).

    tipo ∈ {aprendices, fichas, competencias, historial}
    """
    if request.user.rol != 'administrador':
        return HttpResponseForbidden()
    if not HAS_OPENPYXL:
        messages.error(request, 'openpyxl no esta instalado')
        return redirect('reportes_home')

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active

    TITULO_MAP = {
        'aprendices': 'Reporte de Aprendices',
        'fichas': 'Reporte de Fichas',
        'competencias': 'Reporte de Competencias',
        'historial': 'Historial de Evaluaciones',
    }
    titulo = TITULO_MAP.get(tipo, 'Reporte')

    # Encabezado institucional
    ws.merge_cells('A1:J1')
    ws['A1'] = f'SISTEMA DE GESTION DE SUSTENTACIONES SENA — {titulo.upper()}'
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='2E4057')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    hdr_fill = PatternFill('solid', fgColor='4A6FA5')
    hdr_font = Font(bold=True, color='FFFFFF')

    def escribir_encabezado(row, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal='center')
        ws.row_dimensions[row].height = 20
        return row + 1

    if tipo == 'aprendices':
        headers = ['Documento', 'Nombres', 'Apellidos', 'Email', 'Programa',
                    'Ficha', 'GAES', 'Fase', 'Estado', 'Fecha Registro']
        r = escribir_encabezado(2, headers)
        for ap in Aprendiz.objects.select_related('ficha', 'gaes', 'fase').order_by('nombres'):
            ws.cell(r, 1, ap.documento)
            ws.cell(r, 2, ap.nombres)
            ws.cell(r, 3, ap.apellidos)
            ws.cell(r, 4, ap.email)
            ws.cell(r, 5, ap.programa)
            ws.cell(r, 6, str(ap.ficha) if ap.ficha else '')
            ws.cell(r, 7, str(ap.gaes) if ap.gaes else '')
            ws.cell(r, 8, str(ap.fase) if ap.fase else '')
            ws.cell(r, 9, 'Bloqueado' if ap.bloqueado else 'Activo')
            ws.cell(r, 10, ap.created_at.strftime('%d/%m/%Y') if ap.created_at else '')
            r += 1

    elif tipo == 'fichas':
        headers = ['Numero', 'Programa', 'Jornada', 'GAES', 'Trimestre',
                    'Instructor', 'Estado', 'Cant. Aprendices']
        r = escribir_encabezado(2, headers)
        for f in Ficha.objects.select_related('gaes', 'trimestre', 'instructor').prefetch_related('aprendices').order_by('numero'):
            ws.cell(r, 1, f.numero)
            ws.cell(r, 2, f.programa)
            ws.cell(r, 3, f.jornada)
            ws.cell(r, 4, str(f.gaes) if f.gaes else '')
            ws.cell(r, 5, str(f.trimestre) if f.trimeste else '')
            ws.cell(r, 6, f.instructor.get_full_name() if f.instructor else '')
            ws.cell(r, 7, f.get_estado_display())
            ws.cell(r, 8, f.aprendices.count())
            r += 1

    elif tipo == 'competencias':
        headers = ['Codigo', 'Nombre', 'Fase', 'GAES', 'Ficha', 'Trimestre', 'Activo']
        r = escribir_encabezado(2, headers)
        for comp in Competencia.objects.select_related('fase', 'gaes', 'ficha', 'trimestre').order_by('codigo'):
            ws.cell(r, 1, comp.codigo)
            ws.cell(r, 2, comp.nombre)
            ws.cell(r, 3, str(comp.fase) if comp.fase else '')
            ws.cell(r, 4, str(comp.gaes) if comp.gaes else '')
            ws.cell(r, 5, str(comp.ficha) if comp.ficha else '')
            ws.cell(r, 6, str(comp.trimestre) if comp.trimestre else '')
            ws.cell(r, 7, 'Si' if comp.activo else 'No')
            r += 1

    else:  # historial
        headers = ['Fecha', 'Aprendiz', 'Documento', 'Jurado', 'Checklist', 'Estado', 'Calificacion']
        r = escribir_encabezado(2, headers)
        for ev in Evaluacion.objects.select_related('aprendiz', 'juror', 'checklist').order_by('-fecha'):
            ws.cell(r, 1, ev.fecha.strftime('%d/%m/%Y %H:%M') if ev.fecha else '')
            ws.cell(r, 2, f'{ev.aprendiz.nombres} {ev.aprendiz.apellidos}')
            ws.cell(r, 3, ev.aprendiz.documento)
            ws.cell(r, 4, str(ev.juror.username or ''))
            ws.cell(r, 5, ev.checklist.titulo if ev.checklist else '')
            ws.cell(r, 6, ev.get_estado_display())
            ws.cell(r, 7, str(ev.calificacion_total))
            r += 1

    # Ajuste de anchos
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = min(max_len + 2, 60)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'reporte_{tipo}_{timezone.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
